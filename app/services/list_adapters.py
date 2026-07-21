"""
Adaptateurs de listes officielles : téléchargement + analyse.

Chaque adaptateur produit des enregistrements normalisés consommés par
`list_ingest.ingest`. Le format de chaque source a été établi sur le fichier
réel publié, pas supposé.
"""
from __future__ import annotations

import csv
import hashlib
import logging
import os
import re
import tempfile
import time
import xml.etree.ElementTree as ET
from typing import Iterator

import httpx

logger = logging.getLogger("simandou.list_adapters")

DOWNLOAD_TIMEOUT = 300.0


def _download_to_file(url: str, cache_hours: float = 0) -> str:
    """
    Télécharge en FLUX vers un fichier temporaire.

    Les listes officielles pèsent plusieurs dizaines de Mo (47 Mo pour le
    Royaume-Uni). Les charger entièrement en mémoire faisait tomber l'instance
    (502). On écrit sur disque et on analyse en lecture séquentielle.
    """
    # L'EXTENSION compte : openpyxl refuse un classeur dont le nom n'a pas de
    # suffixe reconnu (.xlsx/.xlsm). Un suffixe « .csv » codé en dur faisait
    # échouer toute source Excel avec un message trompeur sur le format.
    suffix = os.path.splitext(url.split("?")[0])[1] or ".dat"

    # Cache disque : un import volumineux se fait par tranches successives, et
    # sans cache chaque tranche retéléchargerait la source entière — 70 Mo pour
    # l'archive ICIJ, soit des dizaines de gigaoctets sur un import complet.
    if cache_hours > 0:
        digest = hashlib.sha256(url.encode()).hexdigest()[:20]
        cached = os.path.join(tempfile.gettempdir(), f"listcache_{digest}{suffix}")
        if os.path.exists(cached):
            age_h = (time.time() - os.path.getmtime(cached)) / 3600
            if age_h < cache_hours and os.path.getsize(cached) > 1024:
                logger.info("list_download_cache_hit", extra={"source_url": url})
                return cached

    fd, path = tempfile.mkstemp(suffix=suffix, prefix="sanctions_")
    os.close(fd)
    # Plusieurs portails officiels sont protégés par un pare-feu applicatif qui
    # refuse les clients s'annonçant comme des scripts (« python-httpx/… »),
    # surtout depuis une adresse de centre de données. On s'identifie donc comme
    # un navigateur, et on borne le délai de connexion pour qu'un blocage
    # échoue vite et lisiblement au lieu de rester suspendu.
    headers = {
        "User-Agent": ("Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
                       "(KHTML, like Gecko) Chrome/126.0 Safari/537.36"),
        "Accept": "*/*",
        "Accept-Language": "fr,en;q=0.8",
    }
    timeout = httpx.Timeout(DOWNLOAD_TIMEOUT, connect=20.0)
    with httpx.Client(timeout=timeout, follow_redirects=True, headers=headers) as client:
        with client.stream("GET", url) as r:
            r.raise_for_status()
            with open(path, "wb") as f:
                for chunk in r.iter_bytes(chunk_size=1 << 20):
                    f.write(chunk)
    if cache_hours > 0:
        try:
            os.replace(path, cached)
            path = cached
        except OSError:
            pass

    size = os.path.getsize(path)
    if size < 1024:                    # une page d'erreur pèse quelques centaines d'octets
        os.unlink(path)
        raise RuntimeError(
            f"Téléchargement suspect depuis {url} : {size} octets reçus. "
            "La source a probablement refusé la requête."
        )
    return path


# --------------------------------------------------------------------------
# Royaume-Uni — UK Sanctions List (FCDO)
# --------------------------------------------------------------------------
# ATTENTION : la liste consolidée OFSI a été FERMÉE le 28 janvier 2026 et n'est
# plus mise à jour. La UK Sanctions List publiée par le FCDO est désormais
# l'unique source des désignations britanniques.
#
# Particularités du fichier :
#   - une ligne de préambule « Report Date: ... » précède l'en-tête ;
#   - PLUSIEURS lignes par désignation, à regrouper par « Unique ID » ;
#   - pour une personne : Name 1..5 = prénoms, Name 6 = patronyme ;
#     pour une entité ou un navire : Name 6 = dénomination complète ;
#   - la casse de « Name type » est incohérente (Primary Name / Primary name /
#     Primary name variation) : il faut normaliser sous peine de perdre le nom
#     principal d'une partie des désignations.
UK_SANCTIONS_URL = "https://sanctionslist.fcdo.gov.uk/docs/UK-Sanctions-List.csv"
UK_SOURCE_CODE = "UK"
UK_SOURCE_NAME = "Sanctions — Royaume-Uni (UK Sanctions List)"

_NAME_COLS = [f"Name {i}" for i in range(1, 7)]


def _uk_full_name(row: dict) -> str:
    if (row.get("Designation Type") or "").strip().lower() == "individual":
        parts = [(row.get(c) or "").strip() for c in _NAME_COLS]
        return " ".join(p for p in parts if p)
    return (row.get("Name 6") or "").strip()


def fetch_uk_sanctions(url: str = UK_SANCTIONS_URL) -> Iterator[dict]:
    """Télécharge et regroupe la UK Sanctions List par désignation."""
    path = _download_to_file(url)
    grouped: dict[str, dict] = {}
    try:
        # Lecture LIGNE À LIGNE : matérialiser les 57 000 lignes en mémoire
        # coûtait plusieurs centaines de Mo. Seul le regroupement est conservé.
        with open(path, encoding="utf-8-sig", errors="replace", newline="") as fh:
            fh.readline()               # saute « Report Date: ... »
            for row in csv.DictReader(fh):
                _uk_accumulate(row, grouped)
    finally:
        try:
            os.unlink(path)
        except OSError:
            pass

    for entry in grouped.values():
        # Certaines désignations n'ont qu'une variation : elle devient le nom principal.
        if not entry["primary_name"]:
            if not entry["aliases"]:
                continue
            entry["primary_name"] = entry["aliases"].pop(0)
        entry.pop("_seen", None)
        yield entry


def _uk_accumulate(row: dict, grouped: dict[str, dict]) -> None:
    """Agrège une ligne dans la désignation à laquelle elle appartient."""
    uid = (row.get("Unique ID") or "").strip()
    name = _uk_full_name(row)
    if not uid or not name:
        return

    dtype = (row.get("Designation Type") or "").strip().lower()
    entity_type = "person" if dtype == "individual" else "company"
    is_primary = (row.get("Name type") or "").strip().lower() == "primary name"

    entry = grouped.setdefault(uid, {
            "source_ref": uid,
            "entity_type": entity_type,
            "primary_name": "",
            # Le fichier est dénormalisé sur les ADRESSES : le même nom
            # réapparaît sur des dizaines de lignes (jusqu'à 720 pour une seule
            # désignation). On dédoublonne en conservant l'ordre d'apparition,
            # sans quoi une entité se verrait attribuer des centaines d'alias
            # identiques.
            "_seen": {},
            "aliases": [],
            "program": (row.get("Regime Name") or "").strip() or None,
            "listed_on": None,
            "country": None,
            "summary": (row.get("UK Statement of Reasons") or "").strip()[:2000] or None,
            "raw": {"unique_id": uid, "designation_type": row.get("Designation Type"),
                    "source": row.get("Designation source"),
                    "sanctions": row.get("Sanctions Imposed")},
        })
    if name in entry["_seen"]:
        return
    entry["_seen"][name] = True
    if is_primary and not entry["primary_name"]:
        entry["primary_name"] = name
    elif name != entry["primary_name"]:
        entry["aliases"].append(name)


# --------------------------------------------------------------------------
# Canada — Loi sur les mesures économiques spéciales (SEMA)
# --------------------------------------------------------------------------
# Structure établie sur le fichier réel : 5 684 <record> à plat.
#   - personne : <LastName> + <GivenName> ;
#   - entité ou navire : <EntityOrShip> ;
#   - <Aliases> regroupe les alias en un seul texte ;
#   - <Country> porte le RÉGIME de sanction (« Belarus / Bélarus »), pas la
#     nationalité de la personne ;
#   - aucun identifiant unique : la clé stable est (Country, Schedule, Item),
#     vérifiée unique sur l'intégralité du fichier.
CANADA_URL = (
    "https://www.international.gc.ca/world-monde/assets/office_docs/"
    "international_relations-relations_internationales/sanctions/sema-lmes.xml"
)
CANADA_SOURCE_CODE = "CA"
CANADA_SOURCE_NAME = "Sanctions — Canada (LMES/SEMA)"


def _text(node, tag: str) -> str:
    el = node.find(tag)
    return (el.text or "").strip() if el is not None and el.text else ""


def _split_aliases(raw: str) -> list[str]:
    """Les alias sont dans un texte unique, séparés par « ; », « / » ou une virgule."""
    if not raw:
        return []
    parts = re.split(r"[;/]|,(?=\s*[A-ZÀ-Ý])", raw)
    return [p.strip() for p in parts if p and len(p.strip()) > 2]


def fetch_canada_sanctions(url: str = CANADA_URL) -> Iterator[dict]:
    """Télécharge et analyse la liste canadienne (LMES/SEMA)."""
    path = _download_to_file(url)
    try:
        for _, rec in ET.iterparse(path, events=("end",)):
            if rec.tag != "record":
                continue
            regime = _text(rec, "Country")
            entity = _text(rec, "EntityOrShip")
            last, given = _text(rec, "LastName"), _text(rec, "GivenName")

            if entity:
                name, etype = entity, "company"
            else:
                name = " ".join(p for p in (given, last) if p)
                etype = "person"
            if not name:
                rec.clear()
                continue

            ref = "|".join([regime, _text(rec, "Schedule"), _text(rec, "Item")])
            yield {
                "source_ref": ref,
                "entity_type": etype,
                "primary_name": name,
                "aliases": _split_aliases(_text(rec, "Aliases")),
                "program": regime or None,
                "listed_on": _text(rec, "DateOfListing") or None,
                "country": None,     # <Country> = régime, pas la nationalité
                "summary": None,
                "raw": {
                    "schedule": _text(rec, "Schedule"), "item": _text(rec, "Item"),
                    "dob": _text(rec, "DateOfBirthOrShipBuildDate"),
                    "imo": _text(rec, "ShipIMONumber"),
                },
            }
            rec.clear()          # libère la mémoire au fil de la lecture
    finally:
        try:
            os.unlink(path)
        except OSError:
            pass


# --------------------------------------------------------------------------
# Suisse — SECO, liste consolidée
# --------------------------------------------------------------------------
# Structure établie sur le fichier réel (40 Mo, 8 604 cibles) :
#   <target ssid> -> <individual|entity|object> -> <identity> -> <name>
#   <name name-type="primary-name|alias|formerly-known-as">
#       <name-part order name-part-type="family-name|given-name|whole-name|...">
#           <value>…</value><spelling-variant>…</spelling-variant>
#
# Points d'attention :
#   - le nom se RECOMPOSE en ordonnant les <name-part> par `order` ; la partie
#     « title » est écartée (« Mr », « Général »… ne sont pas des noms) ;
#   - les mêmes noms sont répétés dans plusieurs LANGUES : sans dédoublonnage,
#     chaque cible héritait d'alias redondants ;
#   - les variantes orthographiques (y compris en cyrillique) sont de vrais
#     alias : les conserver améliore le rapprochement ;
#   - le programme se résout via <sanctions-set-id> ; les 65 programmes sont
#     déclarés AVANT les cibles, ce qui permet une lecture en un seul passage.
SECO_URL = ("https://www.sesam.search.admin.ch/sesam-search-web/pages/"
            "downloadXmlGesamtliste.xhtml?lang=en&action=downloadXmlGesamtlisteAction")
SECO_SOURCE_CODE = "SECO"
SECO_SOURCE_NAME = "Sanctions — Suisse (SECO)"

_SECO_SKIP_PARTS = {"title", "suffix"}


def _seco_names(node) -> list[tuple[bool, str]]:
    """Retourne [(est_principal, nom)] pour une cible, variantes comprises."""
    out: list[tuple[bool, str]] = []
    for name in node.iter("name"):
        parts = []
        variants: list[str] = []
        for part in name.findall("name-part"):
            if (part.get("name-part-type") or "") in _SECO_SKIP_PARTS:
                continue
            value = (part.findtext("value") or "").strip()
            if value:
                try:
                    order = int(part.get("order") or 0)
                except ValueError:
                    order = 0
                parts.append((order, value))
            variants.extend(
                v.text.strip() for v in part.findall("spelling-variant") if v.text and v.text.strip()
            )
        full = " ".join(v for _, v in sorted(parts, key=lambda x: x[0]))
        is_primary = (name.get("name-type") or "") == "primary-name"
        if full:
            out.append((is_primary, full))
        out.extend((False, v) for v in variants)
    return out


def fetch_seco_sanctions(url: str = SECO_URL) -> Iterator[dict]:
    """Télécharge et analyse la liste consolidée suisse."""
    path = _download_to_file(url)
    programs: dict[str, str] = {}
    depth = 0
    try:
        # ATTENTION : le document contient 8 470 <target> IMBRIQUÉS dans
        # l'historique (<modification><added|removed>), dont 1 698 sous
        # « removed » — c'est-à-dire des personnes RETIRÉES de la liste.
        # Les ingérer reviendrait à signaler comme sanctionnées des personnes
        # délistées. Seules les cibles de premier niveau (enfants directs de la
        # racine, donc profondeur 2) sont des désignations en vigueur.
        for event, el in ET.iterparse(path, events=("start", "end")):
            if event == "start":
                depth += 1
                continue
            el_depth, depth = depth, depth - 1

            if el.tag == "target" and el_depth != 2:
                continue                      # historique : on ignore
            if el.tag == "sanctions-program":
                names = [p for p in el.findall("program-name") if p.get("lang") == "fre"] \
                    or el.findall("program-name")
                label = (names[0].text or "").strip() if names else ""
                for s in el.findall("sanctions-set"):
                    programs[s.get("ssid") or ""] = label
                el.clear()
                continue
            if el.tag != "target":
                continue

            kind_node = next((el.find(k) for k in ("individual", "entity", "object")
                              if el.find(k) is not None), None)
            if kind_node is None:
                el.clear()
                continue
            etype = "person" if kind_node.tag == "individual" else "company"

            primary = ""
            aliases: list[str] = []
            seen: dict[str, bool] = {}
            for is_primary, value in _seco_names(kind_node):
                if value in seen:                 # même nom répété d'une langue à l'autre
                    continue
                seen[value] = True
                if is_primary and not primary:
                    primary = value
                else:
                    aliases.append(value)
            if not primary:
                if not aliases:
                    el.clear()
                    continue
                primary = aliases.pop(0)

            yield {
                "source_ref": el.get("ssid") or "",
                "entity_type": etype,
                "primary_name": primary,
                "aliases": aliases,
                "program": programs.get((el.findtext("sanctions-set-id") or "").strip()) or None,
                "listed_on": None,
                "country": None,
                "summary": (kind_node.findtext("justification") or "").strip()[:2000] or None,
                "raw": {"ssid": el.get("ssid"), "kind": kind_node.tag},
            }
            el.clear()
    finally:
        try:
            os.unlink(path)
        except OSError:
            pass


# --------------------------------------------------------------------------
# Australie — DFAT, Consolidated List
# --------------------------------------------------------------------------
# Classeur Excel (1,3 Mo), établi sur le fichier réel : 11 046 lignes pour
# 3 838 désignations.
#
# La clé de regroupement est le PRÉFIXE NUMÉRIQUE de la référence : « 3080 »
# porte le nom principal, « 3080a » … « 3080ae » ses 32 alias. Sans ce
# regroupement, chaque alias serait importé comme une entité distincte —
# 11 046 entités au lieu de 3 838, et autant de faux positifs autonomes.
# Contrôle de cohérence : le nombre de lignes « Primary Name » (3 838) est
# exactement égal au nombre de références de base.
DFAT_URL = "https://www.dfat.gov.au/sites/default/files/Australian_Sanctions_Consolidated_List.xlsx"
DFAT_SOURCE_CODE = "DFAT"
DFAT_SOURCE_NAME = "Sanctions — Australie (DFAT)"

_DFAT_REF = re.compile(r"^(\d+)")


def fetch_dfat_sanctions(url: str = DFAT_URL) -> Iterator[dict]:
    """Télécharge puis analyse la Consolidated List australienne."""
    path = _download_to_file(url)
    try:
        yield from parse_dfat_sanctions(path)
    finally:
        try:
            os.unlink(path)
        except OSError:
            pass


def parse_dfat_sanctions(path: str) -> Iterator[dict]:
    """
    Analyse un classeur DFAT DÉJÀ présent sur le disque.

    Séparé du téléchargement car le portail australien refuse les requêtes
    venant d'un serveur (connexion acceptée puis maintenue sans réponse). La
    Conformité dépose alors le classeur obtenu depuis son navigateur.
    """
    import openpyxl

    grouped: dict[str, dict] = {}
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header = [str(h or "").strip() for h in next(rows)]

        # Un import qui ne remonte RIEN sans lever d'erreur est le pire des cas
        # en conformité : on croit le contrôle fait alors qu'il ne l'est pas.
        # Si la source change de format, on veut un message, pas un silence.
        required = {"Reference", "Name of Individual or Entity", "Type", "Name Type"}
        missing = required - set(header)
        if missing:
            raise RuntimeError(
                "Colonnes absentes du classeur DFAT : "
                + ", ".join(sorted(missing))
                + f" — colonnes lues : {header[:8]}"
            )

        def cell(row, key: str) -> str:
            try:
                v = row[header.index(key)]
            except (ValueError, IndexError):
                return ""
            return str(v).strip() if v is not None else ""

        for row in rows:
            ref = cell(row, "Reference")
            name = cell(row, "Name of Individual or Entity")
            if not ref or not name:
                continue
            m = _DFAT_REF.match(ref)
            base = m.group(1) if m else ref

            kind = cell(row, "Type").lower()
            etype = "person" if kind == "individual" else "company"
            is_primary = cell(row, "Name Type").lower() == "primary name"

            entry = grouped.setdefault(base, {
                "source_ref": base,
                "entity_type": etype,
                "primary_name": "",
                "aliases": [],
                "_seen": {},
                "program": cell(row, "Committees") or None,
                "listed_on": None,
                "country": cell(row, "Citizenship")[:64] or None,
                "summary": cell(row, "Listing Information")[:2000] or None,
                "raw": {"reference": base, "type": cell(row, "Type"),
                        "dob": cell(row, "Date of Birth")[:120]},
            })
            if name in entry["_seen"]:
                continue
            entry["_seen"][name] = True
            if is_primary and not entry["primary_name"]:
                entry["primary_name"] = name
            elif name != entry["primary_name"]:
                entry["aliases"].append(name)
        wb.close()
    except StopIteration:
        raise RuntimeError("Classeur DFAT vide : aucune ligne d'en-tête.")

    for entry in grouped.values():
        if not entry["primary_name"]:
            if not entry["aliases"]:
                continue
            entry["primary_name"] = entry["aliases"].pop(0)
        entry.pop("_seen", None)
        yield entry


# --------------------------------------------------------------------------
# Guinée — Journal Officiel (Secrétariat Général du Gouvernement)
# --------------------------------------------------------------------------
# Source la plus stratégique pour la BCRG : c'est là que sont publiées les
# nominations aux fonctions publiques, donc les personnes politiquement
# exposées guinéennes — que AUCUNE liste internationale ne couvre.
#
# Établi sur les éditions réelles :
#   - URL prévisible /JO/{année}/guinee-jo-{année}-{n}.pdf, stable de 2022 à
#     2026 ; on ne dépend donc pas de la mise en page du site ;
#   - les PDF portent une COUCHE TEXTE (127 000 caractères pour 36 pages) :
#     aucune reconnaissance optique nécessaire, contrairement aux documents
#     numérisés du dépôt ;
#   - publication bimensuelle, ~24 éditions par an, ~120 personnes par édition.
#
# Extraction : les actes suivent la forme « Décret D/2026/165/PRG/SGG du …,
# portant <objet> », et les personnes « Monsieur|Madame <Prénoms> <NOM> ».
# On rattache chaque personne à l'acte qui la précède, et on QUALIFIE l'acte
# (nomination, décoration, autre) : figurer au Journal Officiel ne fait pas de
# quelqu'un une PPE — la Conformité doit pouvoir distinguer une nomination
# d'une remise de décoration.
GUINEA_JO_SOURCE_CODE = "SGG_GN"
GUINEA_JO_SOURCE_NAME = "PPE Guinée — Journal Officiel (SGG)"
GUINEA_JO_URL = "https://journal-officiel.sgg.gov.gn/JO/{year}/guinee-jo-{year}-{edition:02d}.pdf"

_JO_ACT = re.compile(
    r"(D[ée]cret|DECRET|Arr[êe]t[ée]|ARRETE)\s+([A-ZÀ-Ý]?/?\s?\d{4}\s?/\s?\d+[^\s,]*)"
    r"[^,]{0,60},?\s*portant\s+([^.]{5,160})",
    re.I,
)
_JO_NAME = re.compile(
    r"(?:Monsieur|Madame|M\.|Mme)\s+((?:[A-ZÀ-Ý][\w'’\-]+\s+){0,4}[A-ZÀ-Ý]{2,}[\w'’\-]*)"
)


def _jo_act_kind(objet: str) -> str:
    o = objet.lower()
    if "nomination" in o or "nomme" in o or "nommé" in o:
        return "NOMINATION"
    if "mérite" in o or "merite" in o or "ordre national" in o or "décoration" in o:
        return "DECORATION"
    return "AUTRE"


def parse_guinea_jo(path: str, edition_ref: str = "") -> Iterator[dict]:
    """Extrait les personnes nommément citées d'une édition du Journal Officiel."""
    import pypdf

    reader = pypdf.PdfReader(path)
    raw = "\n".join((p.extract_text() or "") for p in reader.pages)

    # 1) Recolle les mots coupés en fin de ligne : le PDF produit « Secré -
    #    tariats », et un patronyme ainsi scindé serait extrait tronqué ou
    #    manqué. À faire AVANT toute normalisation des espaces.
    txt = re.sub(r"(\w)\s*-\s*\n\s*(\w)", r"\1\2", raw)
    # 2) Normalisation indispensable : les en-têtes d'actes sont coupés par les
    #    sauts de ligne, et l'objet du décret s'en trouve tronqué.
    txt = re.sub(r"\s+", " ", txt)

    acts = [(m.start(), m.group(2).replace(" ", ""), m.group(3).strip())
            for m in _JO_ACT.finditer(txt)]
    if not acts:
        return
    bounds = [p for p, _, _ in acts] + [len(txt)]

    seen: set[tuple[str, str]] = set()
    for i, (_, act_ref, objet) in enumerate(acts):
        block = txt[bounds[i]:bounds[i + 1]]
        kind = _jo_act_kind(objet)
        for name in _JO_NAME.findall(block):
            name = name.strip()
            if len(name.split()) < 2:          # un seul mot : trop ambigu
                continue
            key = (act_ref, name.upper())
            if key in seen:
                continue
            seen.add(key)
            yield {
                "source_ref": f"{edition_ref}|{act_ref}|{name.upper()}",
                "entity_type": "person",
                "primary_name": name,
                "aliases": [],
                # Le « programme » porte la nature de l'acte : c'est lui qui
                # permet à la Conformité de distinguer une nomination d'une
                # décoration au moment de traiter une correspondance.
                "program": f"{kind} — {objet[:120]}",
                "listed_on": None,
                "country": "GN",
                "summary": f"Journal Officiel {edition_ref} · {act_ref} · {objet[:400]}",
                "raw": {"edition": edition_ref, "act_ref": act_ref,
                        "act_kind": kind, "objet": objet[:400]},
            }


def fetch_guinea_jo(year: int = 2026, edition: int = 1) -> Iterator[dict]:
    """Télécharge puis analyse une édition du Journal Officiel guinéen."""
    url = GUINEA_JO_URL.format(year=int(year), edition=int(edition))
    path = _download_to_file(url)
    try:
        yield from parse_guinea_jo(path, edition_ref=f"{year}-{int(edition):02d}")
    finally:
        try:
            os.unlink(path)
        except OSError:
            pass


# --------------------------------------------------------------------------
# Registre des adaptateurs
# --------------------------------------------------------------------------
# ─── ONU / OFAC / UE ──────────────────────────────────────────────────────────
#
# Ces trois listes étaient alimentées par un agent autonome dont la stratégie
# — purge puis rechargement — échouait dès qu'une vérification avait rapproché
# une entité (contrainte screening_matches_entity_id_fkey). Les analyseurs sont
# repris tels quels, éprouvés sur les XML réels ; seul le chargement change.

UN_SOURCE_CODE = "UN"
OFAC_SOURCE_CODE = "OFAC"
EU_SOURCE_CODE = "EU"


def _sanction_entities_to_records(entities) -> Iterator[dict]:
    """Convertit les objets des analyseurs au format du moteur d'ingestion."""
    for e in entities:
        etype = "person" if str(getattr(e.entity_type, "value", e.entity_type)) == "individual" else "company"
        primary = (e.primary_name or "").strip()
        if not primary:
            continue
        aliases = [n.name_raw for n in (e.names or [])
                   if not n.is_primary and (n.name_raw or "").strip()]
        yield {
            # L'identifiant d'origine porte l'idempotence ET la mise à jour :
            # c'est lui qui permet de retrouver une entité d'un run à l'autre
            # au lieu de la recréer sous un nouvel identifiant.
            "source_ref": str(e.source_id),
            "primary_name": primary,
            "entity_type": etype,
            "country": e.country_focus,
            "aliases": aliases,
            "program": ", ".join(e.programs or [])[:255] or None,
            "summary": (e.remarks or None),
            "raw": {"programs": e.programs or [], "source": e.source},
        }


def fetch_un_sanctions() -> Iterator[dict]:
    from app.services.sanctions_sources import un
    yield from _sanction_entities_to_records(un.fetch())


def fetch_ofac_sanctions() -> Iterator[dict]:
    from app.services.sanctions_sources import ofac
    yield from _sanction_entities_to_records(ofac.fetch())


def fetch_eu_sanctions() -> Iterator[dict]:
    from app.services.sanctions_sources import eu
    yield from _sanction_entities_to_records(eu.fetch())


ADAPTERS: dict[str, dict] = {
    UN_SOURCE_CODE: {
        "name": "Nations Unies — Liste consolidée du Conseil de sécurité",
        "fetch": fetch_un_sanctions,
        "record_type": "SANCTION",
        "url": "https://scsanctions.un.org/resources/xml/en/consolidated.xml",
        "label": "ONU — Liste consolidée",
    },
    OFAC_SOURCE_CODE: {
        "name": "États-Unis — OFAC, Specially Designated Nationals",
        "fetch": fetch_ofac_sanctions,
        "record_type": "SANCTION",
        "url": "https://www.treasury.gov/ofac/downloads/sdn.xml",
        "label": "États-Unis — OFAC SDN",
    },
    EU_SOURCE_CODE: {
        "name": "Union européenne — Liste consolidée des sanctions financières",
        "fetch": fetch_eu_sanctions,
        "record_type": "SANCTION",
        "url": "https://webgate.ec.europa.eu/fsd/fsf/public/files/xmlFullSanctionsList_1_1/content?token=dG9rZW4tMjAxNw",
        "label": "Union européenne — FSF",
    },
    UK_SOURCE_CODE: {
        "name": UK_SOURCE_NAME,
        "fetch": fetch_uk_sanctions,
        "record_type": "SANCTION",
        "url": UK_SANCTIONS_URL,
        "label": "Royaume-Uni — UK Sanctions List (FCDO)",
    },
    CANADA_SOURCE_CODE: {
        "name": CANADA_SOURCE_NAME,
        "fetch": fetch_canada_sanctions,
        "record_type": "SANCTION",
        "url": CANADA_URL,
        "label": "Canada — Mesures économiques spéciales (LMES)",
    },
    SECO_SOURCE_CODE: {
        "name": SECO_SOURCE_NAME,
        "fetch": fetch_seco_sanctions,
        "record_type": "SANCTION",
        "url": SECO_URL,
        "label": "Suisse — SECO, liste consolidée",
    },
    GUINEA_JO_SOURCE_CODE: {
        "name": GUINEA_JO_SOURCE_NAME,
        "fetch": fetch_guinea_jo,
        "parse": parse_guinea_jo,
        "record_type": "PEP",
        # Être cité au Journal Officiel n'est pas être sanctionné : le risque
        # attaché reste modéré, à charge pour le scoring de le pondérer.
        "risk_level": "MEDIUM",
        "source_type": "PEP_RULES",
        "accepts": ["year", "edition"],
        "url": "https://journal-officiel.sgg.gov.gn",
        "label": "Guinée — Journal Officiel (nominations et actes)",
    },
    DFAT_SOURCE_CODE: {
        "name": DFAT_SOURCE_NAME,
        "fetch": fetch_dfat_sanctions,
        "parse": parse_dfat_sanctions,
        "upload_only": True,   # le portail refuse les requêtes serveur
        "file_hint": "Classeur Australian_Sanctions_Consolidated_List.xlsx",
        "record_type": "SANCTION",
        "url": DFAT_URL,
        "label": "Australie — DFAT, Consolidated List",
    },
}
